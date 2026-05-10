"""
CLI 命令行接口模块

支持 AI 工具直接调用，提供结构化输出（JSON）和标准退出码。

使用示例：
    # 基础运行（标准输出）
    python cli.py

    # 指定工作目录
    python cli.py --base-dir "C:\\Users\\77188\\Desktop\\期货库存"

    # AI 友好模式：JSON 输出 + 退出码
    python cli.py --json

    # 预览模式（不保存文件）
    python cli.py --dry-run --json

    # 详细日志
    python cli.py -v
"""

import argparse
import json
import os
import sys
from datetime import datetime

from config import DEFAULT_BASE_DIR, DIR_NAME_KEYWORD
from utils import find_base_dir
from main import find_files, setup_directories, backup_original_file
from excel_handler import ExcelReader, ExcelWriter
from processor import InventoryProcessor
from models import ScheduleRecord


def parse_args(argv=None):
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        prog="HSXFuturesInvMgr",
        description="期货库存数据处理工具 v5.0 - 支持 AI 调用的 CLI 版本",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s                           # 基础运行
  %(prog)s --json                    # AI 友好 JSON 输出
  %(prog)s --dry-run --json          # 预览模式，不保存文件
  %(prog)s -b "C:\\path" -v          # 指定目录并显示详细日志
        """,
    )

    parser.add_argument(
        "-b", "--base-dir",
        type=str,
        default=None,
        help=f"指定工作目录（默认自动查找 '{DIR_NAME_KEYWORD}' 目录）",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="以 JSON 格式输出结果摘要（AI 友好模式）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="预览模式：只比对数据，不保存 Excel 文件",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="显示详细处理日志",
    )
    parser.add_argument(
        "--version",
        action="store_true",
        help="显示版本信息",
    )

    return parser.parse_args(argv)


def format_summary_json(summary, exit_code, error_msg=None):
    """
    将处理结果格式化为 JSON 字符串（AI 友好）
    Args:
        summary: dict 或 None
        exit_code: 0=成功, 1=错误, 2=警告
        error_msg: 错误信息
    Returns:
        JSON 字符串
    """
    result = {
        "tool": "HSXFuturesInvMgr",
        "version": "5.0",
        "timestamp": datetime.now().isoformat(),
        "exit_code": exit_code,
        "status": "error" if exit_code != 0 else "success",
    }

    if error_msg:
        result["error"] = error_msg

    if summary:
        result["data"] = summary

    return json.dumps(result, indent=2, ensure_ascii=False)


def log(message, json_mode=False, verbose=False, is_verbose=False):
    """
    统一日志输出
    - json_mode=True: 只输出最终结果 JSON，中间日志丢弃
    - verbose=False + is_verbose=True: 跳过详细日志
    """
    if json_mode and not is_verbose:
        return
    if is_verbose and not verbose:
        return
    print(message)


def main(argv=None):
    """
    CLI 主入口
    Returns:
        int: 退出码（0=成功, 1=错误, 2=警告）
    """
    args = parse_args(argv)

    # 版本信息
    if args.version:
        print("HSXFuturesInvMgr v5.0")
        return 0

    json_mode = args.json
    verbose = args.verbose
    dry_run = args.dry_run

    # ---------- 1. 环境初始化 ----------
    base_dir = args.base_dir
    if base_dir is None:
        base_dir = find_base_dir(DEFAULT_BASE_DIR, DIR_NAME_KEYWORD)
        if not base_dir:
            msg = "Error: Directory not found"
            if json_mode:
                print(format_summary_json(None, 1, msg))
            else:
                print(msg)
            return 1

    os.chdir(base_dir)
    results_dir = setup_directories(base_dir)

    futures_files, yehui_files = find_files(results_dir, base_dir)

    if not futures_files or not yehui_files:
        msg = "Error: Excel files not found"
        if json_mode:
            print(format_summary_json(None, 1, msg))
        else:
            print(msg)
        return 1

    futures_file = futures_files[0]
    yehui_file = yehui_files[0]

    # ---------- 2. 备份 ----------
    if not dry_run:
        backup_path = os.path.join(results_dir, "期货库存明细_备份.xlsx")
        if backup_original_file(futures_file, backup_path):
            log(f"Backup saved: {os.path.basename(backup_path)}", json_mode, verbose)
    else:
        log("[DRY-RUN] Backup skipped", json_mode, verbose)

    log("=" * 60, json_mode, verbose)
    log("Futures Inventory Data Processing Tool v5.0", json_mode, verbose)
    if dry_run:
        log("[DRY-RUN MODE] No files will be saved", json_mode, verbose)
    log("=" * 60, json_mode, verbose)
    log(f"Futures file: {futures_file}", json_mode, verbose)
    log(f"Yehui file: {yehui_file}", json_mode, verbose)
    log("", json_mode, verbose)

    # ---------- 3. 加载工作簿 ----------
    try:
        from openpyxl import load_workbook
        futures_wb = load_workbook(futures_file)
        yehui_wb = load_workbook(yehui_file)
    except Exception as e:
        msg = f"Error: Failed to load Excel files - {e}"
        if json_mode:
            print(format_summary_json(None, 1, msg))
        else:
            print(msg)
            print("提示：请确保文件未被 Excel 占用并已关闭")
        return 1

    reader = ExcelReader(futures_wb)
    writer = ExcelWriter(futures_wb)
    processor = InventoryProcessor()

    # ---------- 4. 读取模板表头 ----------
    template_sheet, template_headers = reader.find_template_sheet()
    if template_sheet is None:
        msg = "Error: No valid template sheet found!"
        if json_mode:
            print(format_summary_json(None, 1, msg))
        else:
            print(msg)
        return 1

    log(f"Template sheet: {template_sheet.title}", json_mode, verbose, is_verbose=True)
    log(f"Standard column count: {len(template_headers)}", json_mode, verbose, is_verbose=True)

    # ---------- 5. 读取数据 ----------
    from config import SCHEDULE_SHEET_NAME, SKIP_SHEET_NAMES

    schedule_data, schedule_coils = reader.read_schedule_data(SCHEDULE_SHEET_NAME)
    log(f"Read {len(schedule_data)} orders from futures schedule", json_mode, verbose)

    yehui_ws = yehui_wb[yehui_wb.sheetnames[0]]
    yehui_records, yehui_coil_info = reader.read_yehui_data(yehui_ws)
    log(f"Loaded {len(yehui_records)} records from Yehui inventory", json_mode, verbose)

    original_all_rows = []
    for sheet_name in futures_wb.sheetnames:
        if sheet_name in SKIP_SHEET_NAMES:
            continue
        ws = futures_wb[sheet_name]
        rows, _ = reader.read_customer_data(ws, sheet_name)
        original_all_rows.extend(rows)
        log(f"  {sheet_name}: {len(rows)} rows", json_mode, verbose, is_verbose=True)

    log(f"Loaded {len(original_all_rows)} rows from original file", json_mode, verbose)

    # ---------- 6. 数据比对 ----------
    preserved, updated, _ = processor.compare_data(original_all_rows, yehui_coil_info, schedule_data)
    original_coil_set = processor.build_coil_set(original_all_rows)
    new_coils = processor.find_new_coils(schedule_data, yehui_records, original_coil_set)

    log(f"Preserved rows:                 {len(preserved)}", json_mode, verbose)
    log(f"Updated coils (warehouse diff): {len(updated)}", json_mode, verbose)
    log(f"New coils from schedule+Yehui:  {len(new_coils)}", json_mode, verbose)

    # ---------- 7. 重建 Sheet ----------
    if not dry_run:
        # 删除旧客户 Sheet
        sheets_to_remove = [s for s in futures_wb.sheetnames if s not in {SCHEDULE_SHEET_NAME, "Sheet1"}]
        for sheet_name in sheets_to_remove:
            del futures_wb[sheet_name]

        from utils import format_date, safe_sheet_name
        from datetime import datetime

        customer_groups = processor.group_by_customer(preserved, updated, new_coils, schedule_data)
        modification_date = format_date(datetime.now())

        for customer_name, items in customer_groups.items():
            safe_name = safe_sheet_name(customer_name)
            ws = writer.create_customer_sheet(safe_name)

            current_row = 2
            prev_order_no = None

            for item_type, item_data in items:
                order_no = item_data.get("order_no", "") if isinstance(item_data, dict) else item_data.order_no

                if prev_order_no is not None and prev_order_no != order_no:
                    current_row += 1

                if item_type == "preserved":
                    writer.write_preserved_row(ws, current_row, item_data.to_dict())
                elif item_type == "updated":
                    writer.write_updated_row(ws, current_row, item_data, modification_date)
                elif item_type == "new":
                    coil_no = item_data["coil_no"]
                    order_no_new = item_data["order_no"]
                    schedule_info = schedule_data.get(order_no_new, ScheduleRecord(order_no="", customer=""))
                    coil_info = yehui_coil_info.get(coil_no, {})
                    writer.write_new_row(ws, current_row, item_data, schedule_info.to_dict(), coil_info, modification_date)

                current_row += 1
                prev_order_no = order_no

            log(f"Created sheet: {safe_name} ({current_row - 2} data rows)", json_mode, verbose, is_verbose=True)

        # 更新期货排程颜色
        schedule_ws = futures_wb[SCHEDULE_SHEET_NAME]
        order_has_yehui = processor.build_order_has_yehui(yehui_records)
        order_in_customer_sheet = processor.build_order_in_customer_sheet(original_all_rows)
        matched_count, unmatched_count = writer.update_schedule_colors(
            schedule_ws, order_has_yehui, order_in_customer_sheet
        )

        log("Schedule color updated:", json_mode, verbose)
        log(f"  Col 3 Green (matched):   {matched_count}", json_mode, verbose)
        log(f"  Col 3 Red (unmatched):   {unmatched_count}", json_mode, verbose)

        # 保存
        futures_wb.save(futures_file)
        log(f"Updated file: {futures_file}", json_mode, verbose)
    else:
        # dry-run: 只计算颜色统计，不实际写入
        order_has_yehui = processor.build_order_has_yehui(yehui_records)
        order_in_customer_sheet = processor.build_order_in_customer_sheet(original_all_rows)
        matched_count = len(order_has_yehui | order_in_customer_sheet)
        unmatched_count = len(set(schedule_data.keys()) - order_has_yehui - order_in_customer_sheet)
        log("[DRY-RUN] Sheets not saved", json_mode, verbose)

    # ---------- 8. 输出结果 ----------
    summary = {
        "preserved": len(preserved),
        "updated": len(updated),
        "new": len(new_coils),
        "matched": matched_count,
        "unmatched": unmatched_count,
    }

    exit_code = 0
    if unmatched_count > 0:
        exit_code = 2  # 有未匹配订单，返回警告码

    if json_mode:
        print(format_summary_json(summary, exit_code))
    else:
        log("", json_mode, verbose)
        log("=" * 60, json_mode, verbose)
        log("SUCCESS! v5.0", json_mode, verbose)
        log("=" * 60, json_mode, verbose)
        log(f"Sheets: {futures_wb.sheetnames}", json_mode, verbose)
        log("", json_mode, verbose)
        log("Summary:", json_mode, verbose)
        log(f"  - Preserved rows (no change):          {len(preserved)}", json_mode, verbose)
        log(f"  - Updated coils  (warehouse changed):  {len(updated)}  -> orange", json_mode, verbose)
        log(f"  - New coils      (schedule+Yehui):     {len(new_coils)}  -> yellow", json_mode, verbose)
        log(f"  - Schedule matched:                   {matched_count}  -> green", json_mode, verbose)
        log(f"  - Schedule unmatched:                  {unmatched_count}  -> red", json_mode, verbose)
        log("", json_mode, verbose)
        log("Processing complete!", json_mode, verbose)

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
