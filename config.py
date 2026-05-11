"""
配置常量模块
集中管理所有常量、颜色、表头等配置
"""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== 路径配置 ====================
# 默认基础目录：None 表示使用脚本所在目录（当前运行路径），
# 确保在 macOS/Windows 下文件均保存在当前运行路径，而非固定桌面路径。
DEFAULT_BASE_DIR = None
DIR_NAME_KEYWORD = "期货"
RESULTS_SUBDIR = "results"

# ==================== 文件匹配模式 ====================
FUTURES_FILE_PATTERN = "*期货库存明细*.xlsx"
YEHUI_FILE_PATTERN = "*烨辉库存*.xlsx"
BACKUP_FILE_NAME = "期货库存明细_备份.xlsx"

# ==================== Sheet 名称 ====================
SCHEDULE_SHEET_NAME = "期货排程"
SKIP_SHEET_NAMES = {"期货排程", "Not Matched", "Sheet1"}

# ==================== 标准表头（23列 + 修改日期）====================
STANDARD_HEADERS = [
    "合同日期", "客户名称", "訂單編號", "鋼捲編號", "訂單厚度", "訂單寬度",
    "出料實測長度", "重量", "毛重", "鍍層代號", "正面漆/紋路色碼",
    "倉別", "移撥日期", "入庫日期", "出料實測厚度", "出料實測寬度",
    "鋅花", "調質", "塗油", "鈍化處理", "正面膜厚", "背面漆色碼", "背面膜厚",
]
MODIFY_DATE_HEADER = "修改日期"
FULL_HEADERS = STANDARD_HEADERS + [MODIFY_DATE_HEADER]

STANDARD_COL_COUNT = len(STANDARD_HEADERS)  # 23
FULL_COL_COUNT = len(FULL_HEADERS)          # 24

# ==================== 样式定义 ====================
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 行状态颜色
UPDATED_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 橙色 - 仓别更新
NEW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")      # 黄色 - 新增
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")    # 绿色 - 已匹配
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # 红色 - 未匹配
NO_FILL = PatternFill(fill_type=None)

# ==================== 字段名变体映射 ====================
FIELD_VARIANTS = {
    "訂單編號": ["訂單編號", "订单编号", "订单号", "Order No"],
    "鋼捲編號": ["鋼捲編號", "钢卷编号", "钢卷号", "Coil No"],
    "倉別": ["倉別", "仓别", "仓库", "Warehouse"],
    "移撥日期": ["移撥日期", "移拨日期", "调拨日期", "Transfer Date"],
    "入庫日期": ["入庫日期", "入库日期", "入库日", "Entry Date"],
}

# ==================== 关键列名（用于索引查找）====================
KEY_COLUMNS = {
    "warehouse": ["倉別", "仓别"],
    "transfer": ["移撥日期", "移拨日期"],
    "entry": ["入庫日期", "入库日期"],
    "coil": ["鋼捲編號", "钢卷编号", "钢卷号"],
    "order": ["訂單編號", "订单编号"],
    "date": ["合同日期", "日期"],
    "customer": ["客户", "客戶"],
}

# ==================== 期货排程列索引（fallback）====================
SCHEDULE_DEFAULT_COLS = {
    "date": 1,
    "customer": 2,
    "order": 3,
    "coil": 4,
}

# ==================== Sheet名限制 ====================
SHEET_NAME_MAX_LEN = 31
SHEET_NAME_INVALID_CHARS = {"\\", "/", "*", "?", ":", "[", "]"}
