"""
Excel 读写处理模块
封装所有与 openpyxl 相关的操作
"""

from collections import defaultdict
from copy import copy
from typing import Dict, List, Tuple, Optional, Any, Set

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import (
    STANDARD_HEADERS, FULL_HEADERS, STANDARD_COL_COUNT, FULL_COL_COUNT,
    HEADER_FILL, HEADER_FONT, THIN_BORDER,
    UPDATED_FILL, NEW_FILL, GREEN_FILL, RED_FILL, NO_FILL,
    SCHEDULE_SHEET_NAME, SKIP_SHEET_NAMES, KEY_COLUMNS, SCHEDULE_DEFAULT_COLS,
)
from models import CoilRecord, ScheduleRecord, YehuiRecord
from utils import format_date, get_field_by_name


def clone_fill(fill: Any) -> Any:
    """复制 openpyxl 填充样式，避免 StyleProxy 无法直接复用"""
    if not fill:
        return None
    try:
        return copy(fill)
    except Exception:
        return None


def resolve_col_index(header_map: Dict[str, int], target_name: str, variants: List[str], default: int = 0) -> int:
    """
    根据表头映射和变体列表解析列索引
    Args:
        header_map: 字段名 -> 列索引
        target_name: 目标字段名
        variants: 字段名变体列表
        default: 默认值
    Returns:
        列索引（从1开始）
    """
    for name in [target_name] + variants:
        if name in header_map:
            return header_map[name]
    return default


class ExcelReader:
    """Excel 读取器"""

    def __init__(self, workbook):
        self.wb = workbook

    def find_template_sheet(self) -> Tuple[Any, List[Tuple[int, str]]]:
        """
        查找标准模板 Sheet（包含关键字段的客户 Sheet）
        Returns:
            (sheet_obj, [(col_idx, header), ...])
        """
        for sname in self.wb.sheetnames:
            if sname in SKIP_SHEET_NAMES:
                continue
            ws = self.wb[sname]
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col_idx).value
                if val:
                    headers.append((col_idx, str(val).strip()))

            header_names = [h for _, h in headers]
            if "倉別" in header_names and "移撥日期" in header_names and "入庫日期" in header_names:
                return ws, headers

        return None, []

    def read_schedule_data(self, sheet_name: str) -> Tuple[Dict[str, ScheduleRecord], Dict[str, Set[str]]]:
        """
        读取期货排程数据
        Returns:
            (order_no -> ScheduleRecord, order_no -> set of coil_no)
        """
        ws = self.wb[sheet_name]

        # 读取表头映射
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx

        col_order = resolve_col_index(header_map, "訂單號碼", KEY_COLUMNS["order"], SCHEDULE_DEFAULT_COLS["order"])
        col_customer = resolve_col_index(header_map, "客户", KEY_COLUMNS["customer"], SCHEDULE_DEFAULT_COLS["customer"])
        col_date = resolve_col_index(header_map, "合同日期", KEY_COLUMNS["date"], SCHEDULE_DEFAULT_COLS["date"])
        col_coil = resolve_col_index(header_map, "鋼捲編號", KEY_COLUMNS["coil"], SCHEDULE_DEFAULT_COLS["coil"])

        schedule_data: Dict[str, ScheduleRecord] = {}
        schedule_coils: Dict[str, Set[str]] = defaultdict(set)

        for row_idx in range(2, ws.max_row + 1):
            order_no = ws.cell(row=row_idx, column=col_order).value
            customer = ws.cell(row=row_idx, column=col_customer).value
            date = ws.cell(row=row_idx, column=col_date).value
            coil_no = ws.cell(row=row_idx, column=col_coil).value

            if order_no:
                order_no_str = str(order_no).strip()
                if order_no_str not in schedule_data:
                    schedule_data[order_no_str] = ScheduleRecord(
                        order_no=order_no_str,
                        customer=str(customer).strip() if customer else "",
                        date=date,
                        row=row_idx,
                    )
                if coil_no:
                    schedule_coils[order_no_str].add(str(coil_no).strip())

        return schedule_data, schedule_coils

    def read_yehui_data(self, worksheet) -> Tuple[Dict[str, YehuiRecord], Dict[str, Dict[str, Any]]]:
        """
        读取烨辉库存表数据
        Returns:
            (key -> YehuiRecord, coil_no -> {warehouse, transfer_date, entry_date})
        """
        ws = worksheet

        # 读取表头
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx

        records: Dict[str, YehuiRecord] = {}
        coil_info: Dict[str, Dict[str, Any]] = {}

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header:
                    row_data[str(header).strip()] = ws.cell(row=row_idx, column=col_idx).value

            order_no = get_field_by_name(row_data, "訂單編號")
            coil_no = get_field_by_name(row_data, "鋼捲編號")

            if order_no and coil_no:
                order_no_str = str(order_no).strip()
                coil_no_str = str(coil_no).strip()
                key = f"{order_no_str}_{coil_no_str}"

                records[key] = YehuiRecord(
                    order_no=order_no_str,
                    coil_no=coil_no_str,
                    data=row_data,
                )

                warehouse = get_field_by_name(row_data, "倉別")
                transfer_date = get_field_by_name(row_data, "移撥日期")
                entry_date = get_field_by_name(row_data, "入庫日期")
                coil_info[coil_no_str] = {
                    "warehouse": str(warehouse).strip() if warehouse else "",
                    "transfer_date": transfer_date,
                    "entry_date": entry_date,
                }

        return records, coil_info

    def read_customer_data(self, worksheet, sheet_name: str) -> Tuple[List[CoilRecord], List[CoilRecord]]:
        """
        读取单个客户 Sheet 的数据
        Returns:
            (all_rows, non_empty_rows)
        """
        ws = worksheet

        # 查找关键列
        header_map = {}
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx

        col_order = resolve_col_index(header_map, "訂單編號", KEY_COLUMNS["order"], 3)
        col_coil = resolve_col_index(header_map, "鋼捲編號", KEY_COLUMNS["coil"], 4)
        col_warehouse = resolve_col_index(header_map, "倉別", KEY_COLUMNS["warehouse"], 12)
        col_transfer = resolve_col_index(header_map, "移撥日期", KEY_COLUMNS["transfer"], 13)
        col_entry = resolve_col_index(header_map, "入庫日期", KEY_COLUMNS["entry"], 14)

        rows: List[CoilRecord] = []

        for row_idx in range(2, ws.max_row + 1):
            # 读取标准列数 + 修改日期列的数据
            row_data = []
            for col_idx in range(1, FULL_COL_COUNT + 1):
                row_data.append(ws.cell(row=row_idx, column=col_idx).value)

            order_no = ws.cell(row=row_idx, column=col_order).value
            if order_no is None or str(order_no).strip() == "":
                continue

            coil_no = ws.cell(row=row_idx, column=col_coil).value
            warehouse = ws.cell(row=row_idx, column=col_warehouse).value
            transfer = ws.cell(row=row_idx, column=col_transfer).value
            entry = ws.cell(row=row_idx, column=col_entry).value
            fill = clone_fill(ws.cell(row=row_idx, column=4).fill)
            modify_date = ws.cell(row=row_idx, column=FULL_COL_COUNT).value

            record = CoilRecord(
                order_no=str(order_no).strip(),
                coil_no=str(coil_no).strip() if coil_no else "",
                warehouse=str(warehouse).strip() if warehouse else "",
                transfer_date=transfer,
                entry_date=entry,
                customer=sheet_name,
                row_data=row_data,
                fill=fill,
                row_idx=row_idx,
                modify_date=modify_date,
            )
            rows.append(record)

        return rows, rows  # 返回相同列表（已过滤空行）


class ExcelWriter:
    """Excel 写入器"""

    def __init__(self, workbook):
        self.wb = workbook

    def create_customer_sheet(self, customer_name: str):
        """创建客户 Sheet 并写入表头"""
        ws = self.wb.create_sheet(title=customer_name)

        for col_idx, header in enumerate(FULL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        return ws

    def write_preserved_row(self, ws, row_idx: int, row_info: Dict[str, Any]):
        """写入保留行（原样保留颜色和修改日期）"""
        row_data = row_info.get("row_data", [])
        fill = row_info.get("fill")
        modify_date = row_info.get("modify_date")

        for col_idx in range(1, FULL_COL_COUNT + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx - 1 < len(row_data):
                cell.value = row_data[col_idx - 1]
            else:
                cell.value = ""
            cell.border = THIN_BORDER
            if fill and hasattr(fill, "fgColor") and fill.fgColor:
                try:
                    cell.fill = clone_fill(fill)
                except Exception:
                    pass

        # 修改日期列：保留原有日期（若 row_data 已包含则不再覆盖，否则单独写入）
        if not row_data or len(row_data) < FULL_COL_COUNT:
            mod_cell = ws.cell(row=row_idx, column=FULL_COL_COUNT)
            if modify_date is not None:
                mod_cell.value = modify_date
            mod_cell.border = THIN_BORDER
            if fill and hasattr(fill, "fgColor") and fill.fgColor:
                try:
                    mod_cell.fill = clone_fill(fill)
                except Exception:
                    pass

    def write_updated_row(self, ws, row_idx: int, update_info: Dict[str, Any], modification_date: str):
        """写入更新行（橙色 - 仓别变动）"""
        row_data = update_info.get("row_data", [])

        for col_idx, header in enumerate(STANDARD_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if header == "倉別":
                cell.value = update_info.get("new_warehouse", "")
            elif header == "移撥日期":
                new_transfer = update_info.get("new_transfer_date")
                cell.value = format_date(new_transfer) if new_transfer else ""
            elif header == "入庫日期":
                new_entry = update_info.get("new_entry_date")
                cell.value = format_date(new_entry) if new_entry else ""
            elif col_idx <= STANDARD_COL_COUNT and col_idx - 1 < len(row_data):
                cell.value = row_data[col_idx - 1]
            else:
                cell.value = ""

            cell.border = THIN_BORDER
            cell.fill = UPDATED_FILL

        # 修改日期列
        mod_cell = ws.cell(row=row_idx, column=FULL_COL_COUNT)
        mod_cell.value = modification_date
        mod_cell.border = THIN_BORDER
        mod_cell.fill = UPDATED_FILL

    def write_new_row(self, ws, row_idx: int, new_info: Dict[str, Any],
                      schedule_info: Dict[str, Any], coil_info: Dict[str, Any],
                      modification_date: str):
        """写入新增行（黄色 - 新增钢卷）"""
        yehui_record = new_info.get("data", {})
        coil_no = new_info["coil_no"]
        order_no = new_info["order_no"]
        contract_date = schedule_info.get("date")
        customer_name = schedule_info.get("customer", "")

        for col_idx, header in enumerate(STANDARD_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if header == "合同日期":
                cell.value = format_date(contract_date) if contract_date else ""
            elif header == "客户名称":
                cell.value = customer_name
            elif header == "訂單編號":
                cell.value = order_no
            elif header == "鋼捲編號":
                cell.value = coil_no
            elif header == "倉別":
                cell.value = coil_info.get("warehouse", "")
            elif header == "移撥日期":
                cell.value = format_date(coil_info.get("transfer_date")) if coil_info.get("transfer_date") else ""
            elif header == "入庫日期":
                cell.value = format_date(coil_info.get("entry_date")) if coil_info.get("entry_date") else ""
            else:
                cell.value = get_field_by_name(yehui_record, header) or ""

            cell.border = THIN_BORDER
            cell.fill = NEW_FILL

        # 修改日期列
        mod_cell = ws.cell(row=row_idx, column=FULL_COL_COUNT)
        mod_cell.value = modification_date
        mod_cell.border = THIN_BORDER
        mod_cell.fill = NEW_FILL

    def update_schedule_colors(self, schedule_ws,
                               order_has_yehui: Set[str],
                               order_in_customer_sheet: Set[str]) -> Tuple[int, int]:
        """
        更新期货排程订单号颜色标注
        Returns:
            (matched_count, unmatched_count)
        """
        matched_count = 0
        unmatched_count = 0

        for row_idx in range(2, schedule_ws.max_row + 1):
            order_no = schedule_ws.cell(row=row_idx, column=3).value

            if order_no is None or str(order_no).strip() == "":
                continue

            order_no_str = str(order_no).strip()
            if order_no_str == "訂單號碼":
                continue

            order_cell = schedule_ws.cell(row=row_idx, column=3)
            spec_cell = schedule_ws.cell(row=row_idx, column=4)
            spec_cell.fill = NO_FILL

            in_yehui = order_no_str in order_has_yehui
            in_customer = order_no_str in order_in_customer_sheet

            if in_yehui or in_customer:
                order_cell.fill = GREEN_FILL
                matched_count += 1
            else:
                order_cell.fill = RED_FILL
                unmatched_count += 1

        return matched_count, unmatched_count
