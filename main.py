"""
期货库存数据处理工具 v5.0 (TDD 重构版)
基于模块化设计，使用 TDD 模式开发

功能：
1. 参考表头，按原顺序排列
2. 合同日期为第一列
3. 日期格式: YYYY/M/D
4. 相同订单号的钢卷放在一起，不同订单号之间空一行
5. 已存在的钢卷：核对仓别，如有变动更新仓别和移拨日期（橙色）
6. 直接更新到 results 文件夹中的原始期货库存明细.xlsx
7. 以期货排程订单为基准：检查每个订单在烨辉表中对应的钢卷
8. 只增加不删减：原文件已有的钢卷即使烨辉表不存在也保留
9. 程序启动前自动备份期货库存明细（覆盖模式，只保留一份备份）
10. 烨辉库存表改用字段名作为唯一ID进行字段匹配（不再依赖列号）
11. 清理多余的修改日期列，只保留唯一的修改日期列
12. 期货排程颜色标注 - 绿色表示订单在烨辉表或客户Sheet中有匹配，红色表示无匹配
"""

import glob
import os
import shutil
from datetime import datetime

from openpyxl import load_workbook

from config import (
    DEFAULT_BASE_DIR, DIR_NAME_KEYWORD, RESULTS_SUBDIR,
    FUTURES_FILE_PATTERN, YEHUI_FILE_PATTERN, BACKUP_FILE_NAME,
    SCHEDULE_SHEET_NAME, SKIP_SHEET_NAMES,
)
from utils import format_date, safe_sheet_name, find_base_dir, is_temp_or_backup_file
from excel_handler import ExcelReader, ExcelWriter
from processor import InventoryProcessor


def find_files(results_dir, current_dir):
    """查找期货库存明细和烨辉库存表文件"""
    futures_files = [
        f for f in glob.glob(os.path.join(results_dir, FUTURES_FILE_PATTERN))
        if not is_temp_or_backup_file(f)
    ]
    yehui_files = [
        f for f in glob.glob(os.path.join(results_dir, YEHUI_FILE_PATTERN))
        if not is_temp_or_backup_file(f)
    ]

    # 如果 results 中没有，从当前目录查找
    if not futures_files:
        futures_files = [
            f for f in glob.glob(os.path.join(current_dir, FUTURES_FILE_PATTERN))
            if not is_temp_or_backup_file(f)
        ]
    if not yehui_files:
        yehui_files = [
            f for f in glob.glob(os.path.join(current_dir, YEHUI_FILE_PATTERN))
            if not is_temp_or_backup_file(f)
        ]

    return futures_files, yehui_files


def backup_original_file(source_path, backup_path):
    """备份原始文件（覆盖模式）"""
    if os.path.exists(source_path):
        shutil.copy2(source_path, backup_path)
        return True
    return False


def setup_directories(base_dir):
    """设置工作目录和 results 目录"""
    results_dir = os.path.join(base_dir, RESULTS_SUBDIR)
    os.makedirs(results_dir, exist_ok=True)
    return results_dir


def main(base_dir=None):
    """主处理函数"""

    # ---------- 1. 环境初始化 ----------
    if base_dir is None:
        base_dir = find_base_dir(DEFAULT_BASE_DIR, DIR_NAME_KEYWORD)
        # 默认路径找不到时，回退到当前脚本所在目录
        if not base_dir:
            base_dir = os.path.dirname(os.path.abspath(__file__))

    os.chdir(base_dir)
    results_dir = setup_directories(base_dir)

    futures_files, yehui_files = find_files(results_dir, base_dir)

    if not futures_files or not yehui_files:
        print("Error: Excel files not found in " + results_dir + " or " + base_dir)
        print("提示：请确认以下文件存在")
        print("  - results/期货库存明细.xlsx")
        print("  - results/烨辉库存表.xlsx")
        return 1

    futures_file = futures_files[0]
    yehui_file = yehui_files[0]

    # ---------- 2. 备份 ----------
    print()
    print("Backing up original file...")
    backup_path = os.path.join(results_dir, BACKUP_FILE_NAME)
    if backup_original_file(futures_file, backup_path):
        print(f"  Backup saved: {os.path.basename(backup_path)}")
    else:
        print("  Warning: Original file not found, skipping backup")

    print("=" * 60)
    print("Futures Inventory Data Processing Tool v5.0")
    print("=" * 60)
    print(f"Futures file: {futures_file}")
    print(f"Yehui file: {yehui_file}")
    print()

    # ---------- 3. 加载工作簿 ----------
    print("Loading workbooks...")
    futures_wb = load_workbook(futures_file)
    yehui_wb = load_workbook(yehui_file)

    reader = ExcelReader(futures_wb)
    writer = ExcelWriter(futures_wb)
    processor = InventoryProcessor()

    # ---------- 4. 读取模板表头 ----------
    print()
    template_sheet, template_headers = reader.find_template_sheet()
    if template_sheet is None:
        print("Error: No valid template sheet found!")
        return 1

    print(f"  Template sheet: {template_sheet.title}")
    print(f"  Standard column count: {len(template_headers)}")

    # ---------- 5. 读取期货排程 ----------
    print()
    print("Step 2: Reading futures schedule data...")
    schedule_data, schedule_coils = reader.read_schedule_data(SCHEDULE_SHEET_NAME)
    print(f"  Read {len(schedule_data)} orders from futures schedule")

    # ---------- 6. 读取烨辉库存表 ----------
    print()
    print("Step 3: Reading Yehui inventory data (by field name)...")
    yehui_ws = yehui_wb[yehui_wb.sheetnames[0]]
    yehui_records, yehui_coil_info = reader.read_yehui_data(yehui_ws)
    print(f"  Loaded {len(yehui_records)} records from Yehui inventory")
    print(f"  Loaded {len(yehui_coil_info)} coil warehouse info")

    # ---------- 7. 读取原文件客户数据 ----------
    print()
    print("Step 3.5: Loading all customer data from original file...")
    original_all_rows: list = []
    original_customer_data: dict = {}

    for sheet_name in futures_wb.sheetnames:
        if sheet_name in SKIP_SHEET_NAMES:
            continue

        ws = futures_wb[sheet_name]
        rows, _ = reader.read_customer_data(ws, sheet_name)
        original_all_rows.extend(rows)
        if rows:
            original_customer_data[sheet_name] = rows
            print(f"    {sheet_name}: {len(rows)} rows")

    print(f"  Loaded {len(original_all_rows)} rows from original file")
    print(f"  Customers: {list(original_customer_data.keys())}")

    # ---------- 8. 数据比对 ----------
    print()
    print("Step 4: Processing data (schedule-based, add only, no deletion)...")

    preserved, updated, _ = processor.compare_data(original_all_rows, yehui_coil_info, schedule_data)
    original_coil_set = processor.build_coil_set(original_all_rows)
    new_coils = processor.find_new_coils(schedule_data, yehui_records, original_coil_set)

    print(f"  Preserved rows:                 {len(preserved)}")
    print(f"  Updated coils (warehouse diff): {len(updated)}")
    print(f"  New coils from schedule+Yehui:  {len(new_coils)}")

    # ---------- 9. 重建客户 Sheet ----------
    print()
    print("Step 5: Rebuilding customer sheets...")

    # 删除旧客户 Sheet（保留期货排程和 Sheet1）
    sheets_to_remove = [s for s in futures_wb.sheetnames if s not in {SCHEDULE_SHEET_NAME, "Sheet1"}]
    for sheet_name in sheets_to_remove:
        del futures_wb[sheet_name]

    # 按客户分组
    customer_groups = processor.group_by_customer(preserved, updated, new_coils, schedule_data)
    modification_date = format_date(datetime.now())

    for customer_name, items in customer_groups.items():
        safe_name = safe_sheet_name(customer_name)
        ws = writer.create_customer_sheet(safe_name)

        current_row = 2
        prev_order_no = None

        for item_type, item_data in items:
            order_no = item_data.get("order_no", "") if isinstance(item_data, dict) else item_data.order_no

            # 不同订单号之间空一行
            if prev_order_no is not None and prev_order_no != order_no:
                current_row += 1

            if item_type == "preserved":
                writer.write_preserved_row(ws, current_row, item_data.to_dict())
            elif item_type == "updated":
                writer.write_updated_row(ws, current_row, item_data, modification_date)
            elif item_type == "new":
                coil_no = item_data["coil_no"]
                order_no_new = item_data["order_no"]
                schedule_info = schedule_data.get(order_no_new, ScheduleRecord(order_no="", customer=""))
                coil_info = yehui_coil_info.get(coil_no, {})
                writer.write_new_row(ws, current_row, item_data, schedule_info.to_dict(), coil_info, modification_date)

            current_row += 1
            prev_order_no = order_no

        print(f"  Created sheet: {safe_name} ({current_row - 2} data rows)")

    # ---------- 10. 更新期货排程颜色 ----------
    print()
    print("Step 6: Updating futures schedule order number column color markers...")

    schedule_ws = futures_wb[SCHEDULE_SHEET_NAME]
    order_has_yehui = processor.build_order_has_yehui(yehui_records)
    order_in_customer_sheet = processor.build_order_in_customer_sheet(original_all_rows)

    matched_count, unmatched_count = writer.update_schedule_colors(
        schedule_ws, order_has_yehui, order_in_customer_sheet
    )

    print(f"  Schedule color updated:")
    print(f"    Col 3 Green (matched):   {matched_count}")
    print(f"    Col 3 Red (unmatched):   {unmatched_count}")

    # ---------- 11. 保存 ----------
    print()
    print("Step 7: Saving updated workbook...")
    futures_wb.save(futures_file)
    print(f"  Updated file: {futures_file}")

    print()
    print("=" * 60)
    print("SUCCESS! v5.0")
    print("=" * 60)
    print(f"Sheets: {futures_wb.sheetnames}")
    print()
    print("Summary:")
    print(f"  - Preserved rows (no change):          {len(preserved)}")
    print(f"  - Updated coils  (warehouse changed):  {len(updated)}  -> orange")
    print(f"  - New coils      (schedule+Yehui):     {len(new_coils)}  -> yellow")
    print(f"  - Schedule matched:                   {matched_count}  -> green")
    print(f"  - Schedule unmatched:                  {unmatched_count}  -> red")
    print()
    print("Logic: schedule orders -> Yehui coil lookup (by field name) -> coil_no dedup -> append only")
    print("Processing complete!")

    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
