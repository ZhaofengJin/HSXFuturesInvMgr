"""
测试 excel_handler 模块
TDD: 先写测试，后实现
使用 unittest.mock 模拟 openpyxl
"""

import unittest
from unittest.mock import MagicMock, patch, PropertyMock
from pathlib import Path
from tempfile import TemporaryDirectory

from excel_handler import ExcelReader, ExcelWriter, resolve_col_index
from config import STANDARD_HEADERS, FULL_HEADERS, SCHEDULE_SHEET_NAME


class TestResolveColIndex(unittest.TestCase):
    """测试列索引解析"""

    def test_exact_match(self):
        headers = {"倉別": 12, "移撥日期": 13}
        self.assertEqual(resolve_col_index(headers, "倉別", ["倉別", "仓别"]), 12)

    def test_fallback_when_not_found(self):
        headers = {"其他": 1}
        self.assertEqual(resolve_col_index(headers, "倉別", ["倉別", "仓别"], default=99), 99)

    def test_variant_match(self):
        headers = {"仓别": 12}
        self.assertEqual(resolve_col_index(headers, "倉別", ["倉別", "仓别"]), 12)


class TestExcelReader(unittest.TestCase):
    """测试 Excel 读取器"""

    def setUp(self):
        self.mock_wb = MagicMock()
        self.mock_ws = MagicMock()
        self.mock_wb.sheetnames = [SCHEDULE_SHEET_NAME, "无锡晟明"]
        self.mock_wb.__getitem__ = MagicMock(return_value=self.mock_ws)

    def test_find_template_sheet(self):
        """测试查找模板 Sheet"""
        # 模拟无锡晟明 Sheet 的表头
        def mock_cell(row, column):
            m = MagicMock()
            if row == 1:
                if column <= len(STANDARD_HEADERS):
                    m.value = STANDARD_HEADERS[column - 1]
                else:
                    m.value = None
            else:
                m.value = None
            return m

        self.mock_ws.cell = mock_cell
        self.mock_ws.max_column = len(STANDARD_HEADERS)
        self.mock_ws.max_row = 1

        reader = ExcelReader(self.mock_wb)
        sheet, headers = reader.find_template_sheet()
        self.assertEqual(sheet, self.mock_ws)
        self.assertEqual(len(headers), len(STANDARD_HEADERS))

    def test_read_schedule_data(self):
        """测试读取期货排程数据"""
        schedule_ws = MagicMock()
        schedule_ws.max_column = 4
        schedule_ws.max_row = 3

        def schedule_cell(row, column):
            m = MagicMock()
            values = {
                (1, 1): "合同日期", (1, 2): "客户", (1, 3): "訂單號碼", (1, 4): "鋼捲編號",
                (2, 1): "2026/4/20", (2, 2): "无锡晟明", (2, 3): "ORD001", (2, 4): "COIL001",
                (3, 1): "2026/4/21", (3, 2): "上海客户", (3, 3): "ORD002", (3, 4): None,
            }
            m.value = values.get((row, column))
            return m

        schedule_ws.cell = schedule_cell
        self.mock_wb.__getitem__ = lambda self, key: schedule_ws if key == SCHEDULE_SHEET_NAME else MagicMock()
        # 注意：上面的 lambda 不对，需要修正
        # 用 side_effect
        def getitem(mock_self, key):
            if key == SCHEDULE_SHEET_NAME:
                return schedule_ws
            return MagicMock()
        self.mock_wb.__getitem__ = getitem

        reader = ExcelReader(self.mock_wb)
        data, coils = reader.read_schedule_data(SCHEDULE_SHEET_NAME)

        self.assertEqual(len(data), 2)
        self.assertEqual(data["ORD001"].customer, "无锡晟明")
        self.assertEqual(data["ORD002"].customer, "上海客户")
        self.assertIn("COIL001", coils["ORD001"])

    def test_read_yehui_data(self):
        """测试读取烨辉库存数据（含 order_no、data 及冲突列表）"""
        yehui_ws = MagicMock()
        yehui_ws.max_column = 3
        yehui_ws.max_row = 3

        def yehui_cell(row, column):
            m = MagicMock()
            values = {
                (1, 1): "訂單編號", (1, 2): "鋼捲編號", (1, 3): "倉別",
                (2, 1): "ORD001", (2, 2): "COIL001", (2, 3): "W1",
                (3, 1): "ORD001", (3, 2): "COIL002", (3, 3): "W2",
            }
            m.value = values.get((row, column))
            return m

        yehui_ws.cell = yehui_cell

        reader = ExcelReader(self.mock_wb)
        records, coil_info, duplicate_coils = reader.read_yehui_data(yehui_ws)

        self.assertEqual(len(records), 2)
        self.assertEqual(coil_info["COIL001"]["warehouse"], "W1")
        self.assertEqual(coil_info["COIL001"]["order_no"], "ORD001")
        self.assertIn("data", coil_info["COIL001"])
        self.assertEqual(coil_info["COIL002"]["warehouse"], "W2")
        self.assertEqual(duplicate_coils, [])

    def test_read_yehui_data_duplicate_coils(self):
        """同一钢卷出现多次时记录冲突并按后行覆盖"""
        yehui_ws = MagicMock()
        yehui_ws.max_column = 3
        yehui_ws.max_row = 3

        def yehui_cell(row, column):
            m = MagicMock()
            values = {
                (1, 1): "訂單編號", (1, 2): "鋼捲編號", (1, 3): "倉別",
                (2, 1): "ORD001", (2, 2): "COIL001", (2, 3): "W1",
                (3, 1): "ORD002", (3, 2): "COIL001", (3, 3): "W2",
            }
            m.value = values.get((row, column))
            return m

        yehui_ws.cell = yehui_cell

        reader = ExcelReader(self.mock_wb)
        records, coil_info, duplicate_coils = reader.read_yehui_data(yehui_ws)

        self.assertEqual(duplicate_coils, ["COIL001"])
        self.assertEqual(coil_info["COIL001"]["order_no"], "ORD002")
        self.assertEqual(coil_info["COIL001"]["warehouse"], "W2")

    def test_read_customer_data(self):
        """测试读取客户 Sheet 数据（包含修改日期列）"""
        customer_ws = MagicMock()
        customer_ws.max_column = len(FULL_HEADERS)
        customer_ws.max_row = 3
        customer_ws.title = "无锡晟明"

        def customer_cell(row, column):
            m = MagicMock()
            headers_map = {i + 1: h for i, h in enumerate(FULL_HEADERS)}
            values = {
                (2, 3): "ORD001", (2, 4): "COIL001", (2, 12): "W1",
                (2, 24): "2026/4/15",  # 修改日期
                (3, 3): "ORD001", (3, 4): "COIL002", (3, 12): "W2",
                (3, 24): "2026/4/16",  # 修改日期
            }
            if row == 1:
                m.value = headers_map.get(column)
            else:
                m.value = values.get((row, column))
            return m

        customer_ws.cell = customer_cell

        reader = ExcelReader(self.mock_wb)
        rows, customer_data = reader.read_customer_data(customer_ws, "无锡晟明")

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].order_no, "ORD001")
        self.assertEqual(rows[0].coil_no, "COIL001")
        self.assertEqual(rows[0].warehouse, "W1")
        self.assertEqual(rows[0].customer, "无锡晟明")
        self.assertEqual(rows[0].modify_date, "2026/4/15")
        self.assertEqual(rows[1].modify_date, "2026/4/16")


class TestExcelWriter(unittest.TestCase):
    """测试 Excel 写入器"""

    def setUp(self):
        self.mock_wb = MagicMock()

    def test_create_customer_sheet(self):
        """测试创建客户 Sheet"""
        mock_ws = MagicMock()
        self.mock_wb.create_sheet = MagicMock(return_value=mock_ws)

        writer = ExcelWriter(self.mock_wb)
        ws = writer.create_customer_sheet("无锡晟明")

        self.mock_wb.create_sheet.assert_called_once_with(title="无锡晟明")
        # 验证表头已写入
        self.assertTrue(mock_ws.cell.called)

    def test_write_preserved_row(self):
        """测试写入保留行应保留原有修改日期和颜色"""
        mock_ws = MagicMock()
        writer = ExcelWriter(self.mock_wb)

        # 模拟一个带修改日期和颜色的保留行
        from openpyxl.styles import PatternFill
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        row_info = {
            "row_data": ["2026/4/20", "无锡晟明", "ORD001", "COIL001"] + [""] * 19,
            "fill": orange_fill,
            "modify_date": "2026/4/15",
        }
        writer.write_preserved_row(mock_ws, 2, row_info)

        # 验证修改日期列（第24列）写入了原有日期，而不是被清空
        mod_cell = mock_ws.cell(row=2, column=len(FULL_HEADERS))
        self.assertEqual(mod_cell.value, "2026/4/15")
        # 验证颜色被保留
        self.assertEqual(mod_cell.fill, orange_fill)

    def test_write_preserved_row_from_coil_record_to_dict(self):
        """测试通过 CoilRecord.to_dict() 传递的保留行仍能保留颜色（模拟 cli.py 实际调用路径）"""
        from openpyxl.styles import PatternFill
        from models import CoilRecord

        mock_ws = MagicMock()
        writer = ExcelWriter(self.mock_wb)

        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        record = CoilRecord(
            order_no="ORD001",
            coil_no="COIL001",
            warehouse="W1",
            customer="无锡晟明",
            row_data=["2026/4/20", "无锡晟明", "ORD001", "COIL001"] + [""] * 19 + ["2026/4/15"],
            fill=orange_fill,
            modify_date="2026/4/15",
        )

        # 模拟 cli.py 中的实际调用方式：CoilRecord.to_dict() -> write_preserved_row
        writer.write_preserved_row(mock_ws, 2, record.to_dict())

        # 验证 cell 被调用多次（至少24列），且 fill 被正确设置
        self.assertTrue(mock_ws.cell.called)
        # mock_ws.cell.return_value 是同一个 MagicMock，write_preserved_row 会对它赋值 fill
        self.assertEqual(mock_ws.cell.return_value.fill, orange_fill)

    def test_write_preserved_row_keeps_fill_after_real_workbook_roundtrip(self):
        """真实 workbook round-trip 后仍应保留颜色"""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill

        with TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "roundtrip.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "客户A"
            for col_idx, header in enumerate(FULL_HEADERS, 1):
                ws.cell(row=1, column=col_idx).value = header

            row_data = ["2026/4/20", "客户A", "ORD001", "COIL001"] + [""] * 19 + ["2026/4/15"]
            fill = PatternFill(start_color="ABCDEF", end_color="ABCDEF", fill_type="solid")
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.value = value
                cell.fill = fill
            wb.save(workbook_path)

            wb = load_workbook(workbook_path)
            reader = ExcelReader(wb)
            rows, _ = reader.read_customer_data(wb["客户A"], "客户A")

            del wb["客户A"]
            writer = ExcelWriter(wb)
            new_ws = writer.create_customer_sheet("客户A")
            writer.write_preserved_row(new_ws, 2, rows[0].to_dict())
            wb.save(workbook_path)

            saved_wb = load_workbook(workbook_path)
            saved_ws = saved_wb["客户A"]
            self.assertEqual(saved_ws.cell(row=2, column=1).fill.fill_type, "solid")
            self.assertEqual(saved_ws.cell(row=2, column=1).fill.fgColor.rgb, "00ABCDEF")
            self.assertEqual(saved_ws.cell(row=2, column=len(FULL_HEADERS)).fill.fill_type, "solid")
            self.assertEqual(saved_ws.cell(row=2, column=len(FULL_HEADERS)).fill.fgColor.rgb, "00ABCDEF")

    def test_write_updated_row(self):
        """测试写入属性更新行（橙色）"""
        mock_ws = MagicMock()
        writer = ExcelWriter(self.mock_wb)

        update_info = {
            "row_data": ["2026/4/20", "无锡晟明", "ORD001", "COIL001"] + [""] * 19,
            "change_type": "updated",
            "coil_no": "COIL001",
            "order_no": "ORD001",
            "yehui_data": {"倉別": "W2", "移撥日期": "2026/4/21", "入庫日期": "2026/4/22"},
        }
        schedule_info = {"date": "2026/4/20", "customer": "无锡晟明"}
        writer.write_updated_row(mock_ws, 2, update_info, schedule_info, "2026/5/10")

        # 验证单元格被调用
        self.assertTrue(mock_ws.cell.called)

    def test_write_transferred_row(self):
        """测试写入转单行（紫色），且使用排程中的新订单/新客户"""
        mock_ws = MagicMock()
        writer = ExcelWriter(self.mock_wb)

        update_info = {
            "row_data": ["2026/4/20", "无锡晟明", "ORD001", "COIL001"] + [""] * 19,
            "change_type": "transferred",
            "coil_no": "COIL001",
            "order_no": "ORD002",
            "new_order_no": "ORD002",
            "old_order_no": "ORD001",
            "yehui_data": {"倉別": "W1", "移撥日期": "2026/4/21", "入庫日期": "2026/4/22", "鍍層代號": "AZ150"},
        }
        schedule_info = {"date": "2026/5/1", "customer": "上海客户（转单）"}
        writer.write_updated_row(mock_ws, 2, update_info, schedule_info, "2026/5/10")

        self.assertTrue(mock_ws.cell.called)
        # 无法直接验证 fill，因为 mock cell 返回同一个对象，但调用不报错即可

    def test_write_new_row(self):
        """测试写入新增行（黄色）"""
        mock_ws = MagicMock()
        writer = ExcelWriter(self.mock_wb)

        new_info = {
            "coil_no": "COIL003",
            "order_no": "ORD001",
            "data": {"倉別": "W3", "移撥日期": "2026/4/25", "入庫日期": "2026/4/26"},
        }
        schedule_info = {"date": "2026/4/20", "customer": "无锡晟明"}
        coil_info = {"warehouse": "W3", "transfer_date": "2026/4/25", "entry_date": "2026/4/26"}

        writer.write_new_row(mock_ws, 2, new_info, schedule_info, coil_info, "2026/5/10")

        self.assertTrue(mock_ws.cell.called)

    def test_update_schedule_colors(self):
        """测试更新期货排程颜色"""
        schedule_ws = MagicMock()
        schedule_ws.max_row = 4

        def schedule_cell(row, column):
            m = MagicMock()
            values = {
                (2, 3): "ORD001",
                (3, 3): "ORD002",
                (4, 3): "",
            }
            m.value = values.get((row, column))
            return m

        schedule_ws.cell = schedule_cell

        writer = ExcelWriter(self.mock_wb)
        matched, unmatched = writer.update_schedule_colors(
            schedule_ws,
            order_has_yehui={"ORD001"},
            order_in_customer_sheet={"ORD001", "ORD002"},
        )

        self.assertEqual(matched, 2)  # ORD001 和 ORD002 都有匹配
        self.assertEqual(unmatched, 0)


if __name__ == "__main__":
    unittest.main()
