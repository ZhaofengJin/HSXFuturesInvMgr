"""
期货库存数据处理工具 v4.5
功能：
1. 参考表头，按原顺序排列
2. 合同日期为第一列
3. 日期格式: 2026/4/20
4. 相同订单号的钢卷放在一起，不同订单号之间空一行
5. 已存在的钢卷：核对仓别，如有变动更新仓别和移拨日期（橙色）
6. 直接更新到results文件夹中的原始期货库存明细.xlsx
7. [v4.2] 以期货排程订单为基准：检查每个订单在烨辉表中对应的钢卷
         根据钢卷号判断是否已存在于对应客户Sheet，不存在则新增（黄色）
8. [v4.2] 只增加不删减：原文件已有的钢卷即使烨辉表不存在也保留
9. [v4.3] 程序启动前自动备份期货库存明细（覆盖模式，只保留一份备份）
10. [v4.4] 修复：限制只读取标准列数数据，避免多余的"修改日期"列导致数据错位
11. [v4.5] 修复：烨辉库存表改用字段名作为唯一ID进行字段匹配（不再依赖列号）
12. [v4.5] 修复：清理多余的修改日期列，只保留Col 24这唯一的修改日期列
13. [v4.5] 修复：期货排程颜色标注 - 绿色表示订单在烨辉表或客户Sheet中有匹配，红色表示两处都无匹配
14. [v4.5] 修复：产品规格列（Col 4）清除特殊颜色，改为与其他列一样
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict
import glob
import os
from datetime import datetime


def format_date(dt):
    """格式化日期为 2026/4/20 格式"""
    if dt is None:
        return ''
    if isinstance(dt, datetime):
        return f"{dt.year}/{dt.month}/{dt.day}"
    if isinstance(dt, str):
        try:
            dt_obj = datetime.strptime(dt.split()[0], '%Y-%m-%d')
            return f"{dt_obj.year}/{dt_obj.month}/{dt_obj.day}"
        except:
            return dt
    return str(dt)


def get_yehui_field_by_name(yehui_row_data, field_name):
    """
    从烨辉表行数据中通过字段名获取值
    烨辉表的字段名可能略有差异，需要处理常见变体
    """
    if field_name in yehui_row_data:
        return yehui_row_data[field_name]

    # 处理常见的字段名变体
    field_variants = {
        '訂單編號': ['訂單編號', '订单编号', '订单号', 'Order No'],
        '鋼捲編號': ['鋼捲編號', '钢卷编号', '钢卷号', 'Coil No'],
        '倉別': ['倉別', '仓别', '仓库', 'Warehouse'],
        '移撥日期': ['移撥日期', '移拨日期', '调拨日期', 'Transfer Date'],
        '入庫日期': ['入庫日期', '入库日期', '入库日', 'Entry Date'],
    }

    variants = field_variants.get(field_name, [field_name])
    for variant in variants:
        if variant in yehui_row_data:
            return yehui_row_data[variant]

    return None


def load_previous_output(results_dir):
    """加载上次运行的结果，提取已有的订单-钢卷对和仓别信息"""
    import glob

    # 查找原始期货库存明细文件（排除备份文件）
    futures_files = [f for f in glob.glob(os.path.join(results_dir, "*期货库存明细*.xlsx"))
                     if not os.path.basename(f).startswith('~$')
                     and '备份' not in f
                     and '.bak' not in f]

    if not futures_files:
        print("  No previous output found")
        return None, {}

    # 使用找到的文件
    latest_file = futures_files[0]
    print(f"  Loading from: {os.path.basename(latest_file)}")

    try:
        wb = load_workbook(latest_file)
        existing_coils = set()  # (order_no, coil_no) tuples
        coil_details = {}  # coil_no -> {仓别, 移拨日期, 入庫日期}

        # 遍历所有客户Sheet
        for sheet_name in wb.sheetnames:
            if sheet_name in ['期货排程', 'Not Matched', 'Sheet1']:
                continue

            ws = wb[sheet_name]
            # 查找"倉別"列的索引
            warehouse_col = None
            transfer_col = None
            entry_col = None
            coil_col = None
            order_col = None

            for col_idx in range(1, min(ws.max_column + 1, 30)):
                header = ws.cell(row=1, column=col_idx).value
                if header:
                    header_str = str(header).strip()
                    if header_str in ['倉別', '仓别']:
                        warehouse_col = col_idx
                    elif header_str in ['移撥日期', '移拨日期']:
                        transfer_col = col_idx
                    elif header_str in ['入庫日期', '入库日期']:
                        entry_col = col_idx
                    elif header_str in ['鋼捲編號', '钢卷编号', '钢卷号']:
                        coil_col = col_idx
                    elif header_str in ['訂單編號', '订单编号']:
                        order_col = col_idx

            for row_idx in range(2, ws.max_row + 1):
                order_no = ws.cell(row_idx, order_col).value if order_col else None
                coil_no = ws.cell(row_idx, coil_col).value if coil_col else None
                warehouse = ws.cell(row_idx, warehouse_col).value if warehouse_col else None
                transfer_date = ws.cell(row_idx, transfer_col).value if transfer_col else None
                entry_date = ws.cell(row_idx, entry_col).value if entry_col else None

                if order_no and coil_no:
                    existing_coils.add((str(order_no).strip(), str(coil_no).strip()))
                    coil_key = str(coil_no).strip()
                    coil_details[coil_key] = {
                        'warehouse': str(warehouse).strip() if warehouse else '',
                        'transfer_date': transfer_date,
                        'entry_date': entry_date
                    }

        print(f"  Found {len(existing_coils)} existing coils")
        print(f"  Found {len(coil_details)} coil details")
        return existing_coils, coil_details
    except Exception as e:
        print(f"  Error loading previous output: {e}")
        import traceback
        traceback.print_exc()
        return None, {}


def main():
    """主处理函数"""

    # 找到工作目录
    base_dir = None
    for item in os.listdir(r"C:\Users\77188\Desktop"):
        if not item.startswith('~$') and os.path.isdir(os.path.join(r"C:\Users\77188\Desktop", item)):
            if "期货" in item:
                base_dir = os.path.join(r"C:\Users\77188\Desktop", item)
                break

    if not base_dir:
        print("Error: Directory not found")
        return

    os.chdir(base_dir)

    # 创建results文件夹
    results_dir = os.path.join(base_dir, "results")
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    # 查找文件（优先从results文件夹读取，备用从当前目录读取）
    # 先在results文件夹中查找（排除临时文件和备份文件）
    futures_files = [f for f in glob.glob(os.path.join(results_dir, "*期货库存明细*.xlsx"))
                     if not os.path.basename(f).startswith('~$')
                     and '备份' not in f
                     and '.bak' not in f]
    yehui_files = [f for f in glob.glob(os.path.join(results_dir, "*烨辉库存*.xlsx")) if not os.path.basename(f).startswith('~$')]

    # 如果results中没有，从当前目录查找
    if not futures_files:
        futures_files = [f for f in glob.glob("*期货库存明细*.xlsx")
                        if not os.path.basename(f).startswith('~$') and '备份' not in f and '.bak' not in f]
    if not yehui_files:
        yehui_files = [f for f in glob.glob("*烨辉库存*.xlsx") if not os.path.basename(f).startswith('~$')]

    if not futures_files or not yehui_files:
        print("Error: Excel files not found")
        return

    futures_file = futures_files[0] if futures_files else None
    yehui_file = yehui_files[0] if yehui_files else None

    if not futures_file or not yehui_file:
        print("Error: Excel files not found")
        return

    # === 备份原始文件（覆盖模式，只保留一份备份）===
    print()
    print("Backing up original file...")
    import shutil
    backup_file = os.path.join(results_dir, "期货库存明细_备份.xlsx")
    if os.path.exists(futures_file):
        shutil.copy2(futures_file, backup_file)
        print(f"  Backup saved: {os.path.basename(backup_file)}")
    else:
        print(f"  Warning: Original file not found, skipping backup")

    print("=" * 60)
    print("Futures Inventory Data Processing Tool v4.5")
    print("=" * 60)
    print(f"Futures file: {futures_file}")
    print(f"Yehui file: {yehui_file}")
    print()

    # 加载工作簿
    print("Loading workbooks...")
    futures_wb = load_workbook(futures_file)
    yehui_wb = load_workbook(yehui_file)

    # === Step 1: 读取表头作为模板 ===
    # 优先使用Sheet2作为标准模板
    template_sheet = None
    template_headers = []
    standard_col_count = 0

    # 查找标准表头sheet（Sheet2或第一个客户sheet）
    for sname in futures_wb.sheetnames:
        if sname in ['期货排程', 'Not Matched', 'Sheet1']:
            continue
        ws = futures_wb[sname]
        # 读取表头
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                headers.append((col_idx, header))

        # 过滤掉"修改日期"列，只保留标准表头
        # 标准表头到"背面膜厚"（第23列）为止
        standard_headers = []
        for col_idx, header in headers:
            header_str = str(header).strip()
            # 跳过多余的"修改日期"列
            if header_str == '修改日期' and col_idx > 24:
                continue
            standard_headers.append((col_idx, header))

        # 判断是否是有效模板（包含倉別、移撥日期等关键字段）
        header_names = [h for _, h in standard_headers]
        if '倉別' in header_names and '移撥日期' in header_names and '入庫日期' in header_names:
            template_sheet = futures_wb[sname]
            template_headers = standard_headers
            print(f"  Using {sname} as template")
            break

    if template_sheet is None:
        print("Error: No valid template sheet found!")
        return

    # 计算标准列数（排除Col 24之后的多余修改日期列）
    standard_col_count = 23  # 默认标准23列（到背面膜厚为止）
    for col_idx, header in template_headers:
        if col_idx > standard_col_count:
            break

    print(f"  Template sheet: {template_sheet.title}")
    print(f"  Standard column count: {standard_col_count}")

    # === Step 2: 读取期货排程数据 ===
    print()
    print("Step 2: Reading futures schedule data...")
    schedule_sheet = futures_wb['期货排程']

    # 读取期货排程的所有数据
    # Col 1=合同日期, Col 2=客户, Col 3=訂單號碼
    schedule_data = {}
    schedule_coils = defaultdict(set)  # 记录每个订单在期货排程中的钢卷号

    # 查找期货排程表头列索引
    schedule_col_map = {}
    for col_idx in range(1, schedule_sheet.max_column + 1):
        header = schedule_sheet.cell(row=1, column=col_idx).value
        if header:
            header_str = str(header).strip()
            if header_str in ['合同日期', '日期']:
                schedule_col_map['date'] = col_idx
            elif header_str in ['客户', '客戶']:
                schedule_col_map['customer'] = col_idx
            elif header_str in ['訂單號碼', '订单号码', '订单编号']:
                schedule_col_map['order'] = col_idx
            elif header_str in ['鋼捲編號', '钢卷编号', '钢卷号']:
                schedule_col_map['coil'] = col_idx

    for row_idx in range(2, schedule_sheet.max_row + 1):
        order_col = schedule_col_map.get('order', 3)
        customer_col = schedule_col_map.get('customer', 2)
        date_col = schedule_col_map.get('date', 1)
        coil_col = schedule_col_map.get('coil', 4)

        order_no = schedule_sheet.cell(row=row_idx, column=order_col).value
        customer = schedule_sheet.cell(row=row_idx, column=customer_col).value
        date = schedule_sheet.cell(row=row_idx, column=date_col).value
        coil_no = schedule_sheet.cell(row=row_idx, column=coil_col).value if coil_col else None

        if order_no:
            order_no_str = str(order_no).strip()
            if order_no_str not in schedule_data:
                schedule_data[order_no_str] = {
                    'row': row_idx,
                    'customer': str(customer).strip() if customer else '',
                    'date': date,
                    'order_no': order_no_str
                }
            if coil_no:
                schedule_coils[order_no_str].add(str(coil_no).strip())

    print(f"  Read {len(schedule_data)} orders from futures schedule")

    # === Step 3: 读取烨辉库存表数据（以字段名作为唯一ID）===
    print()
    print("Step 3: Reading Yehui inventory data (by field name)...")
    yehui_sheet = yehui_wb[yehui_wb.sheetnames[0]]

    # 读取烨辉表全部表头（字段名->列索引映射）
    yehui_header_map = {}  # field_name -> col_idx
    for col_idx in range(1, yehui_sheet.max_column + 1):
        val = yehui_sheet.cell(row=1, column=col_idx).value
        if val:
            yehui_header_map[str(val).strip()] = col_idx

    print(f"  Yehui header fields: {list(yehui_header_map.keys())}")

    # 读取烨辉表数据，按钢卷号组织（用于仓别核对）
    yehui_data = defaultdict(dict)
    yehui_coil_info = {}  # coil_no -> {warehouse, transfer_date, entry_date}

    for row_idx in range(2, yehui_sheet.max_row + 1):
        # 读取整行数据为字典 {字段名: 值}
        row_data = {}
        for col_idx in range(1, yehui_sheet.max_column + 1):
            header = yehui_sheet.cell(row=1, column=col_idx).value
            if header:
                row_data[str(header).strip()] = yehui_sheet.cell(row=row_idx, column=col_idx).value

        order_no = get_yehui_field_by_name(row_data, '訂單編號')
        coil_no = get_yehui_field_by_name(row_data, '鋼捲編號')

        if order_no and coil_no:
            order_no_str = str(order_no).strip()
            coil_no_str = str(coil_no).strip()

            key = f"{order_no_str}_{coil_no_str}"
            yehui_data[key] = {
                'order_no': order_no_str,
                'coil_no': coil_no_str,
                'data': row_data  # 整行数据，以字段名为key
            }

            # 记录钢卷的仓别信息（通过字段名获取）
            warehouse = get_yehui_field_by_name(row_data, '倉別')
            transfer_date = get_yehui_field_by_name(row_data, '移撥日期')
            entry_date = get_yehui_field_by_name(row_data, '入庫日期')
            yehui_coil_info[coil_no_str] = {
                'warehouse': str(warehouse).strip() if warehouse else '',
                'transfer_date': transfer_date,
                'entry_date': entry_date
            }

    print(f"  Loaded {len(yehui_data)} records from Yehui inventory")
    print(f"  Loaded {len(yehui_coil_info)} coil warehouse info")

    # === Step 3.5: 加载原文件中所有客户数据 ===
    print()
    print("Step 3.5: Loading all customer data from original file...")

    # 定义日志文件
    log_file = os.path.join(base_dir, "processing_log.txt")

    print(f"  Standard column count: {standard_col_count}")

    # 读取原文件中的所有客户sheet数据
    original_customer_data = {}  # customer -> list of rows
    original_all_rows = []  # 所有原始数据行

    for sheet_name in futures_wb.sheetnames:
        if sheet_name in ['期货排程', 'Not Matched', 'Sheet1']:
            continue

        ws = futures_wb[sheet_name]

        # 查找关键列的索引
        col_map = {}
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                header_str = str(header).strip()
                if header_str in ['倉別', '仓别']:
                    col_map['warehouse'] = col_idx
                elif header_str in ['移撥日期', '移拨日期']:
                    col_map['transfer'] = col_idx
                elif header_str in ['入庫日期', '入库日期']:
                    col_map['entry'] = col_idx
                elif header_str in ['鋼捲編號', '钢卷编号', '钢卷号']:
                    col_map['coil'] = col_idx
                elif header_str in ['訂單編號', '订单编号']:
                    col_map['order'] = col_idx

        customer_rows = []

        for row_idx in range(2, ws.max_row + 1):
            # 只读取标准列数的数据，避免读取多余的"修改日期"列
            row_data = []
            for col_idx in range(1, standard_col_count + 1):
                row_data.append(ws.cell(row=row_idx, column=col_idx).value)

            # 跳过空行（订单之间的分隔行）
            order_col = col_map.get('order', 3)
            coil_col = col_map.get('coil', 4)
            warehouse_col = col_map.get('warehouse', 12)
            transfer_col = col_map.get('transfer', 13)
            entry_col = col_map.get('entry', 14)

            order_no = ws.cell(row=row_idx, column=order_col).value
            if order_no is None or str(order_no).strip() == '':
                continue

            coil_no_val = ws.cell(row=row_idx, column=coil_col).value if coil_col else None

            # 记录原始行数据和样式
            warehouse_val = ws.cell(row=row_idx, column=warehouse_col).value if warehouse_col else None
            transfer_val = ws.cell(row=row_idx, column=transfer_col).value if transfer_col else None
            entry_val = ws.cell(row=row_idx, column=entry_col).value if entry_col else None

            row_info = {
                'row_data': row_data,
                'order_no': str(order_no).strip(),
                'coil_no': str(coil_no_val).strip() if coil_no_val else '',
                'warehouse': str(warehouse_val).strip() if warehouse_val else '',
                'transfer_date': transfer_val,
                'entry_date': entry_val,
                'row_idx': row_idx,
                'fill': ws.cell(row=row_idx, column=4).fill if ws.cell(row=row_idx, column=4).fill else None,
                'customer': sheet_name  # 记录所属客户sheet名称
            }
            original_all_rows.append(row_info)
            customer_rows.append(row_info)

        if customer_rows:
            original_customer_data[sheet_name] = customer_rows
            print(f"    {sheet_name}: {len(customer_rows)} rows")
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"    {sheet_name}: {len(customer_rows)} rows\n")

    print(f"  Loaded {len(original_all_rows)} rows from original file")
    print(f"  Customers: {list(original_customer_data.keys())}")
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"  Loaded {len(original_all_rows)} rows from original file\n")
        f.write(f"  Customers: {list(original_customer_data.keys())}\n")

    # === Step 3.6: 构建烨辉表钢卷映射 ===
    print()
    print("Step 3.6: Building Yehui coil mapping...")

    # coil_no -> yehui record 映射
    yehui_coil_map = {}  # coil_no -> full record
    for key, record in yehui_data.items():
        coil_no = record['coil_no']
        yehui_coil_map[coil_no] = record

    print(f"  Built mapping for {len(yehui_coil_map)} coils from Yehui")

    # === Step 4: 处理数据（以期货排程订单为基准，只增加不删减）===
    print()
    print("Step 4: Processing data (schedule-based, add only, no deletion)...")

    new_coils = []       # 需要新增的钢卷（期货排程有订单 + 烨辉表有数据 + 原文件没有）
    updated_coils = []   # 仓别变动的钢卷（橙色）
    preserved_rows = []  # 保留的原有行（无变化）

    # 建立原文件钢卷集合（快速查找，避免重复添加）
    original_coil_set = {row_info['coil_no'] for row_info in original_all_rows if row_info['coil_no']}

    # ── Step 4.1: 处理原文件已有的所有行 ──────────────────────────────
    # 目的：检查仓别是否有变动（如有则橙色标注），无变动则原样保留
    for row_info in original_all_rows:
        coil_no = row_info['coil_no']
        old_warehouse = row_info['warehouse']

        if coil_no in yehui_coil_info:
            # 从 yehui_coil_info 获取最新仓别
            new_warehouse = yehui_coil_info[coil_no].get('warehouse', '')

            if old_warehouse != new_warehouse:
                # 仓别有变动 → 标记为 updated（橙色）
                updated_coils.append({
                    **row_info,
                    'new_warehouse': new_warehouse,
                    'new_transfer_date': yehui_coil_info[coil_no].get('transfer_date'),
                    'new_entry_date': yehui_coil_info[coil_no].get('entry_date')
                })
            else:
                # 仓别无变动 → 原样保留
                preserved_rows.append(row_info)
        else:
            # 烨辉表中不存在该钢卷 → 仍然保留（只增加不删减）
            preserved_rows.append(row_info)

    # ── Step 4.2: 以期货排程订单为基准，找出需要新增的钢卷 ─────────────
    # 遍历期货排程中的每个订单 → 在烨辉表中查找该订单的钢卷
    # → 如果钢卷不在原文件中，则标记为新增（黄色）
    for order_no, schedule_info in schedule_data.items():
        customer = schedule_info.get('customer', '')
        if not customer:
            continue

        for key, yehui_record in yehui_data.items():
            if yehui_record['order_no'] != order_no:
                continue

            coil_no = yehui_record['coil_no']

            # 根据钢卷号检查是否已存在于原文件
            if coil_no not in original_coil_set:
                new_coils.append({
                    'coil_no':  coil_no,
                    'order_no': order_no,
                    'customer': customer,  # 直接记录客户名，后续无需再查
                    'data':     yehui_record['data']
                })

    total_preserved = len(preserved_rows)
    total_updated   = len(updated_coils)
    total_new       = len(new_coils)

    print(f"  Preserved rows:                 {total_preserved}")
    print(f"  Updated coils (warehouse diff): {total_updated}")
    print(f"  New coils from schedule+Yehui:  {total_new}")

    # === Step 5: 清理旧客户Sheet，重建数据 ===
    print()
    print("Step 5: Rebuilding customer sheets...")

    # 删除旧的客户Sheet（保留期货排程和Sheet1）
    sheets_to_remove = [s for s in futures_wb.sheetnames
                       if s not in ['期货排程', 'Sheet1']]
    for sheet_name in sheets_to_remove:
        del futures_wb[sheet_name]

    # 样式定义
    header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    updated_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 橙色 - 仓别更新
    new_fill     = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色 - 新增

    # 准备标准表头（只到Col 23，删除Col 24之后的所有"修改日期"列）
    # 标准表头：合同日期, 客户名称, 訂單編號, 鋼捲編號, 訂單厚度, 訂單寬度, 出料實測長度, 重量, 毛重,
    #           鍍層代號, 正面漆/紋路色碼, 倉別, 移撥日期, 入庫日期, 出料實測厚度, 出料實測寬度,
    #           鋅花, 調質, 塗油, 鈍化處理, 正面膜厚, 背面漆色碼, 背面膜厚
    # Col 24: 修改日期（唯一的修改日期列）
    standard_headers = [
        '合同日期', '客户名称', '訂單編號', '鋼捲編號', '訂單厚度', '訂單寬度',
        '出料實測長度', '重量', '毛重', '鍍層代號', '正面漆/紋路色碼',
        '倉別', '移撥日期', '入庫日期', '出料實測厚度', '出料實測寬度',
        '鋅花', '調質', '塗油', '鈍化處理', '正面膜厚', '背面漆色碼', '背面膜厚'
    ]
    headers = standard_headers + ['修改日期']  # 24列

    # ── 按客户分组数据 ──────────────────────────────────────────────────
    customer_groups = defaultdict(list)

    # 保留行：直接使用原文件中记录的客户名称
    for row_info in preserved_rows:
        customer = row_info.get('customer', '')
        if customer:
            customer_groups[customer].append(('preserved', row_info))

    # 更新行：优先用原文件记录的客户名，再从期货排程补充
    for update_info in updated_coils:
        customer = update_info.get('customer', '')
        if not customer:
            order_no = update_info.get('order_no', '')
            customer = schedule_data.get(order_no, {}).get('customer', '')
        if customer:
            customer_groups[customer].append(('updated', update_info))

    # 新增行：直接使用 Step 4.2 中已记录的客户名（无需再查）
    for new_info in new_coils:
        customer = new_info.get('customer', '')
        if customer:
            customer_groups[customer].append(('new', new_info))

    modification_date = format_date(datetime.now())

    # ── 创建每个客户的 Sheet ────────────────────────────────────────────
    for customer_name, items in customer_groups.items():
        safe_name = customer_name[:31]
        for char in ['\\', '/', '*', '?', ':', '[', ']']:
            safe_name = safe_name.replace(char, '_')

        ws = futures_wb.create_sheet(title=safe_name)

        # 写入表头（只有24列）
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value     = header
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border    = thin_border

        # 写入数据
        current_row  = 2
        prev_order_no = None

        # 按订单号排序（相同订单的钢卷聚合在一起）
        items.sort(key=lambda x: (x[1].get('order_no', ''), 0 if x[0] == 'preserved' else 1))

        for item_type, item_data in items:
            order_no = item_data.get('order_no', '')

            # 不同订单号之间空一行
            if prev_order_no is not None and prev_order_no != order_no:
                current_row += 1

            # ── preserved：原样写回，保留原有填色 ──────────────────────
            if item_type == 'preserved':
                row_data = item_data['row_data']
                fill     = item_data.get('fill')

                # 只写入标准23列的数据
                for col_idx in range(1, standard_col_count + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if col_idx - 1 < len(row_data):
                        cell.value = row_data[col_idx - 1]
                    else:
                        cell.value = ''
                    cell.border = thin_border
                    if fill and hasattr(fill, 'fgColor') and fill.fgColor:
                        try:
                            cell.fill = fill
                        except Exception:
                            pass

                # 修改日期列留空
                ws.cell(row=current_row, column=24).border = thin_border

            # ── updated：仓别有变动，橙色标注 ──────────────────────────
            elif item_type == 'updated':
                row_data = item_data.get('row_data', [])

                # Col 1-11, Col 15-23: 从原行数据获取
                # Col 12(倉別): 使用new_warehouse
                # Col 13(移撥日期): 使用new_transfer_date
                # Col 14(入庫日期): 使用new_entry_date
                for col_idx, header in enumerate(standard_headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx)

                    if header == '倉別':
                        cell.value = item_data.get('new_warehouse', '')
                    elif header == '移撥日期':
                        new_transfer = item_data.get('new_transfer_date')
                        cell.value = format_date(new_transfer) if new_transfer else ''
                    elif header == '入庫日期':
                        new_entry = item_data.get('new_entry_date')
                        cell.value = format_date(new_entry) if new_entry else ''
                    elif col_idx <= standard_col_count and col_idx - 1 < len(row_data):
                        cell.value = row_data[col_idx - 1]
                    else:
                        cell.value = ''

                    cell.border = thin_border
                    cell.fill   = updated_fill

                # 修改日期列（Col 24）
                mod_cell         = ws.cell(row=current_row, column=24)
                mod_cell.value   = modification_date
                mod_cell.border  = thin_border
                mod_cell.fill    = updated_fill

            # ── new：从烨辉表新增，黄色标注 ────────────────────────────
            elif item_type == 'new':
                yehui_record  = item_data.get('data', {})
                coil_no       = item_data['coil_no']
                order_no_new  = item_data['order_no']
                schedule_info = schedule_data.get(order_no_new, {})
                contract_date = schedule_info.get('date')
                coil_info     = yehui_coil_info.get(coil_no, {})  # 从 yehui_coil_info 取仓别

                for col_idx, header in enumerate(standard_headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx)

                    if header == '合同日期':
                        cell.value = format_date(contract_date) if contract_date else ''
                    elif header == '客户名称':
                        cell.value = customer_name
                    elif header == '訂單編號':
                        cell.value = order_no_new
                    elif header == '鋼捲編號':
                        cell.value = coil_no
                    elif header == '倉別':
                        cell.value = coil_info.get('warehouse', '')
                    elif header == '移撥日期':
                        cell.value = format_date(coil_info.get('transfer_date')) if coil_info.get('transfer_date') else ''
                    elif header == '入庫日期':
                        cell.value = format_date(coil_info.get('entry_date')) if coil_info.get('entry_date') else ''
                    else:
                        # 从烨辉表数据中通过字段名获取
                        cell.value = get_yehui_field_by_name(yehui_record, header) or ''

                    cell.border = thin_border
                    cell.fill   = new_fill

                # 修改日期列（Col 24）
                mod_cell        = ws.cell(row=current_row, column=24)
                mod_cell.value  = modification_date
                mod_cell.border = thin_border
                mod_cell.fill   = new_fill

            current_row   += 1
            prev_order_no  = order_no

        print(f"  Created sheet: {safe_name} ({current_row - 2} data rows)")

    # === Step 6: 更新期货排程订单号码列(Col 3)颜色标注 ===
    print()
    print("Step 6: Updating futures schedule order number column (Col 3) color markers...")

    # 构建订单号 -> 是否有烨辉数据 的映射
    order_has_yehui = {}  # order_no -> bool (是否在烨辉表中有对应钢卷)

    # 从 yehui_data 构建：遍历所有烨辉记录，收集订单号
    for key, yehui_record in yehui_data.items():
        order_no = yehui_record['order_no']
        order_has_yehui[order_no] = True

    # 构建订单号 -> 是否在期货库存表客户Sheet中存在 的映射
    order_in_customer_sheet = set()  # 在客户Sheet中存在过的订单号
    for row_info in original_all_rows:
        order_no = row_info.get('order_no', '')
        if order_no:
            order_in_customer_sheet.add(order_no)

    print(f"  Orders in Yehui:    {len(order_has_yehui)}")
    print(f"  Orders in customer sheets: {len(order_in_customer_sheet)}")

    # 定义颜色
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色 - 已匹配
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # 红色 - 未匹配
    no_fill = PatternFill(fill_type=None)  # 无填充（清除特殊颜色）

    # 更新期货排程Col 3（訂單號碼）的颜色
    matched_count = 0
    unmatched_count = 0

    for row_idx in range(2, schedule_sheet.max_row + 1):
        order_no = schedule_sheet.cell(row=row_idx, column=3).value

        # 跳过空行和表头行
        if order_no is None or str(order_no).strip() == '':
            continue
        order_no_str = str(order_no).strip()

        # 跳过表头行
        if order_no_str == '訂單號碼':
            continue

        # 获取订单号码列（Col 3）
        order_cell = schedule_sheet.cell(row=row_idx, column=3)

        # 清除产品规格列（Col 4）的特殊颜色，改为与其他列一样
        spec_cell = schedule_sheet.cell(row=row_idx, column=4)
        spec_cell.fill = no_fill

        # 颜色标注逻辑：
        # 绿色：订单在烨辉表中有匹配 OR 在期货库存表对应客户Sheet中存在
        # 红色：烨辉表不存在 AND 期货库存表对应客户Sheet中也不存在
        in_yehui = order_no_str in order_has_yehui
        in_customer_sheet = order_no_str in order_in_customer_sheet

        if in_yehui or in_customer_sheet:
            order_cell.fill = green_fill
            matched_count += 1
        else:
            order_cell.fill = red_fill
            unmatched_count += 1

    print(f"  Schedule color updated:")
    print(f"    Col 3 Green (matched):   {matched_count}")
    print(f"    Col 3 Red (unmatched):   {unmatched_count}")

    # === Step 7: 保存更新后的文件 ===
    print()
    print("Step 7: Saving updated workbook...")

    futures_wb.save(futures_file)
    print(f"  Updated file: {futures_file}")

    print()
    print("=" * 60)
    print("SUCCESS! v4.5")
    print("=" * 60)
    print(f"Sheets: {futures_wb.sheetnames}")
    print()
    print("Summary:")
    print(f"  - Preserved rows (no change):          {total_preserved}")
    print(f"  - Updated coils  (warehouse changed):  {total_updated}  → orange")
    print(f"  - New coils      (schedule+Yehui):     {total_new}  → yellow")
    print(f"  - Schedule matched:                   {matched_count}  → green")
    print(f"  - Schedule unmatched:                  {unmatched_count}  → red")
    print()
    print("Logic: schedule orders → Yehui coil lookup (by field name) → coil_no dedup → append only")
    print("Processing complete!")


if __name__ == "__main__":
    main()
